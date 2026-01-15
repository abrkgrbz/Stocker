"""
Sample data generator for Stocker migration templates.
Generates 100 sample customers and 100 sample products in Excel format.
"""
import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# ========================================
# Customer Data Generation
# ========================================

# Turkish company name components
COMPANY_PREFIXES = [
    "Anadolu", "Ege", "Marmara", "Karadeniz", "Akdeniz", "Trakya", "İç Anadolu",
    "Doğu", "Güneydoğu", "Batı", "Kuzey", "Merkez", "Global", "Mega", "Süper",
    "Altın", "Gümüş", "Yeşil", "Mavi", "Kırmızı", "Beyaz", "Siyah", "Turuncu",
    "Yıldız", "Güneş", "Ay", "Deniz", "Dağ", "Orman", "Nehir", "Göl"
]

COMPANY_TYPES = [
    "Ticaret", "Sanayi", "İnşaat", "Tekstil", "Gıda", "Otomotiv", "Elektrik",
    "Elektronik", "Mobilya", "Kimya", "Plastik", "Metal", "Makine", "Tarım",
    "Hayvancılık", "Turizm", "Lojistik", "Nakliyat", "Danışmanlık", "Yazılım",
    "Bilişim", "Enerji", "Madencilik", "İlaç", "Kozmetik", "Ambalaj", "Kağıt",
    "Cam", "Seramik", "Deri", "Ayakkabı", "Konfeksiyon", "Halı", "Perde"
]

COMPANY_SUFFIXES = ["A.Ş.", "Ltd. Şti.", "San. Tic. A.Ş.", "San. ve Tic. Ltd. Şti.", "Holding"]

# Turkish cities and districts
CITIES = {
    "İstanbul": ["Kadıköy", "Beşiktaş", "Şişli", "Ümraniye", "Maltepe", "Ataşehir", "Bakırköy", "Beylikdüzü", "Kartal", "Pendik"],
    "Ankara": ["Çankaya", "Keçiören", "Yenimahalle", "Mamak", "Etimesgut", "Sincan", "Altındağ", "Gölbaşı"],
    "İzmir": ["Konak", "Karşıyaka", "Bornova", "Buca", "Çiğli", "Bayraklı", "Gaziemir", "Aliağa"],
    "Bursa": ["Nilüfer", "Osmangazi", "Yıldırım", "Mudanya", "Gemlik", "İnegöl"],
    "Antalya": ["Muratpaşa", "Kepez", "Konyaaltı", "Alanya", "Manavgat", "Serik"],
    "Adana": ["Seyhan", "Çukurova", "Yüreğir", "Sarıçam", "Ceyhan"],
    "Konya": ["Selçuklu", "Meram", "Karatay", "Ereğli", "Akşehir"],
    "Gaziantep": ["Şahinbey", "Şehitkamil", "Nizip", "İslahiye"],
    "Mersin": ["Yenişehir", "Mezitli", "Toroslar", "Akdeniz", "Tarsus"],
    "Kayseri": ["Melikgazi", "Kocasinan", "Talas", "Develi"]
}

STREETS = [
    "Atatürk Cad.", "İstiklal Cad.", "Cumhuriyet Cad.", "Barbaros Bulvarı", "Bağdat Cad.",
    "Ankara Cad.", "İzmir Cad.", "Bursa Cad.", "Fatih Cad.", "Mevlana Cad.",
    "Konya Cad.", "Antalya Cad.", "Fevzi Çakmak Cad.", "Gazi Mustafa Kemal Cad.",
    "Mehmet Akif Cad.", "Yunus Emre Cad.", "Mimar Sinan Cad.", "Kanuni Sultan Süleyman Cad."
]

INDUSTRIES = [
    "Perakende", "Toptan Ticaret", "İmalat", "İnşaat", "Lojistik", "Gıda",
    "Tekstil", "Otomotiv", "Elektronik", "Sağlık", "Eğitim", "Turizm",
    "Bilişim", "Finans", "Enerji", "Tarım", "Madencilik", "Kimya"
]

PAYMENT_TERMS = ["Peşin", "15 Gün", "30 Gün", "45 Gün", "60 Gün", "90 Gün"]

def generate_tax_number():
    """Generate Turkish tax number (10 digits for companies)"""
    return ''.join([str(random.randint(0, 9)) for _ in range(10)])

def generate_phone():
    """Generate Turkish phone number"""
    area_codes = ["212", "216", "312", "232", "224", "242", "322", "342", "324", "352"]
    return f"0{random.choice(area_codes)} {random.randint(100, 999)} {random.randint(10, 99)} {random.randint(10, 99)}"

def generate_email(company_name):
    """Generate email from company name"""
    domain = company_name.lower().replace(" ", "").replace(".", "")[:15]
    # Remove Turkish characters
    tr_chars = {"ı": "i", "ğ": "g", "ü": "u", "ş": "s", "ö": "o", "ç": "c",
                "İ": "i", "Ğ": "g", "Ü": "u", "Ş": "s", "Ö": "o", "Ç": "c"}
    for tr, en in tr_chars.items():
        domain = domain.replace(tr, en)
    return f"info@{domain}.com.tr"

def generate_customers(count=100):
    """Generate sample customer data"""
    customers = []
    used_names = set()

    for i in range(1, count + 1):
        # Generate unique company name
        while True:
            prefix = random.choice(COMPANY_PREFIXES)
            comp_type = random.choice(COMPANY_TYPES)
            suffix = random.choice(COMPANY_SUFFIXES)
            company_name = f"{prefix} {comp_type} {suffix}"
            if company_name not in used_names:
                used_names.add(company_name)
                break

        city = random.choice(list(CITIES.keys()))
        district = random.choice(CITIES[city])
        street = random.choice(STREETS)
        building_no = random.randint(1, 200)

        customer = {
            "Firma/Kişi Adı*": company_name,
            "E-Posta*": generate_email(company_name),
            "Telefon": generate_phone(),
            "Web Sitesi": f"www.{company_name.lower().replace(' ', '').replace('.', '')[:12]}.com.tr".replace("ı", "i").replace("ğ", "g").replace("ü", "u").replace("ş", "s").replace("ö", "o").replace("ç", "c"),
            "Sektör": random.choice(INDUSTRIES),
            "Adres": f"{street} No:{building_no}",
            "İlçe": district,
            "İl": city,
            "Ülke": "Türkiye",
            "Posta Kodu": f"{random.randint(10, 81)}{random.randint(100, 999)}",
            "Vergi/TC No": generate_tax_number(),
            "Vergi Dairesi": f"{city} Vergi Dairesi",
            "Yetkili Kişi": f"{random.choice(['Ahmet', 'Mehmet', 'Ali', 'Mustafa', 'Hasan', 'Hüseyin', 'İbrahim', 'Osman', 'Yusuf', 'Fatma', 'Ayşe', 'Emine', 'Hatice', 'Zeynep'])} {random.choice(['Yılmaz', 'Kaya', 'Demir', 'Çelik', 'Şahin', 'Yıldız', 'Yıldırım', 'Öztürk', 'Aydın', 'Özdemir'])}",
            "Kredi Limiti": random.choice([0, 5000, 10000, 25000, 50000, 100000, 250000, 500000]),
            "Açıklama": f"{random.choice(['Önemli', 'Yeni', 'VIP', 'Potansiyel', 'Aktif'])} müşteri - {random.choice(['düzenli sipariş', 'büyük hacim', 'hızlı ödeme', 'uzun vadeli ortaklık'])} beklentisi"
        }
        customers.append(customer)

    return customers

# ========================================
# Product Data Generation
# ========================================

CATEGORIES = {
    "Elektronik": ["Telefon", "Tablet", "Laptop", "Monitör", "Klavye", "Mouse", "Kulaklık", "Hoparlör", "Kamera", "Yazıcı"],
    "Gıda": ["Un", "Şeker", "Tuz", "Yağ", "Pirinç", "Makarna", "Konserve", "Baharat", "Çay", "Kahve"],
    "Tekstil": ["T-Shirt", "Gömlek", "Pantolon", "Ceket", "Etek", "Elbise", "Kazak", "Mont", "Eşarp", "Kravat"],
    "Mobilya": ["Masa", "Sandalye", "Koltuk", "Dolap", "Yatak", "Sehpa", "Kitaplık", "Gardırop", "Komodin", "TV Ünitesi"],
    "Temizlik": ["Deterjan", "Yumuşatıcı", "Çamaşır Suyu", "Bulaşık Deterjanı", "Cam Temizleyici", "Yer Temizleyici", "Sünger", "Bez", "Eldiven", "Fırça"],
    "Kozmetik": ["Şampuan", "Krem", "Parfüm", "Ruj", "Fondöten", "Maskara", "Oje", "Deodorant", "Diş Macunu", "Sabun"],
    "Ofis": ["Kalem", "Defter", "Dosya", "Zımba", "Makas", "Cetvel", "Silgi", "Kalemtıraş", "Klasör", "Yapışkanlı Not"],
    "Otomotiv": ["Motor Yağı", "Antifriz", "Fren Balatasısı", "Filtre", "Bujı", "Akü", "Lastik", "Far", "Silecek", "Cam Suyu"],
    "İnşaat": ["Çimento", "Tuğla", "Demir", "Kum", "Çakıl", "Alçı", "Boya", "Fayans", "Seramik", "PVC Boru"],
    "Bahçe": ["Tohum", "Gübre", "Saksı", "Toprak", "Hortum", "Makas", "Kürek", "Tırmık", "Çim Biçme", "Sulama Sistemi"]
}

BRANDS = {
    "Elektronik": ["Samsung", "Apple", "LG", "Sony", "Asus", "HP", "Dell", "Lenovo", "Xiaomi", "Huawei"],
    "Gıda": ["Ülker", "Eti", "Nestle", "Unilever", "Pınar", "Sütaş", "Torku", "Bizim", "Komili", "Yudum"],
    "Tekstil": ["LC Waikiki", "Koton", "Defacto", "Mavi", "Colin's", "Network", "Vakko", "Beymen", "Kiğılı", "Damat"],
    "Mobilya": ["İstikbal", "Bellona", "Doğtaş", "Çilek", "Kelebek", "Mondi", "Alfemo", "Yataş", "Vivense", "Mudo"],
    "Temizlik": ["Ariel", "Persil", "OMO", "Domestos", "Cif", "Fairy", "Mr. Muscle", "Ace", "Vernel", "Yumoş"],
    "Kozmetik": ["L'Oreal", "Nivea", "Dove", "Garnier", "Pantene", "Head&Shoulders", "Axe", "Colgate", "Signal", "Oral-B"],
    "Ofis": ["Faber-Castell", "Pelikan", "Pilot", "Bic", "Pentel", "Uni", "Schneider", "Stabilo", "Staedtler", "Rotring"],
    "Otomotiv": ["Castrol", "Mobil", "Shell", "Total", "Liqui Moly", "Bosch", "Valeo", "Denso", "NGK", "Champion"],
    "İnşaat": ["Çimsa", "Akçansa", "Marshall", "Filli Boya", "Dyo", "Polisan", "Vitra", "Kaleseramik", "Eczacıbaşı", "Ulusoy"],
    "Bahçe": ["Scotts", "Miracle-Gro", "Bayer", "Gardena", "Husqvarna", "Stihl", "Bosch Garden", "Black+Decker", "Makita", "Hitachi"]
}

UNITS = ["Adet", "Kg", "Lt", "Mt", "M2", "Paket", "Kutu", "Koli", "Set", "Çift"]

PRODUCT_TYPES = ["Mamül", "Hammadde", "Yarı Mamül", "Sarf Malzeme"]

def generate_barcode():
    """Generate EAN-13 barcode"""
    # Turkish prefix 869
    return f"869{random.randint(1000000000, 9999999999)}"

def generate_sku(category, index):
    """Generate SKU code"""
    cat_code = category[:3].upper().replace("İ", "I").replace("Ş", "S").replace("Ü", "U").replace("Ö", "O").replace("Ç", "C").replace("Ğ", "G")
    return f"{cat_code}-{str(index).zfill(4)}"

def generate_products(count=100):
    """Generate sample product data"""
    products = []
    used_codes = set()

    categories_list = list(CATEGORIES.keys())

    for i in range(1, count + 1):
        category = random.choice(categories_list)
        product_base = random.choice(CATEGORIES[category])
        brand = random.choice(BRANDS[category])

        # Generate unique product code
        while True:
            code = f"URN-{str(random.randint(10000, 99999))}"
            if code not in used_codes:
                used_codes.add(code)
                break

        # Generate variant suffix
        variants = ["", " Pro", " Plus", " Max", " Mini", " Lite", " Premium", " Standart", " Ekonomik"]
        sizes = ["", " (S)", " (M)", " (L)", " (XL)", " (100ml)", " (250ml)", " (500ml)", " (1L)", " (5L)"]
        colors = ["", " Beyaz", " Siyah", " Mavi", " Kırmızı", " Yeşil", " Gri", " Bej"]

        product_name = f"{brand} {product_base}{random.choice(variants)}{random.choice(sizes + colors)}"

        # Base price based on category
        base_prices = {
            "Elektronik": (100, 15000),
            "Gıda": (5, 500),
            "Tekstil": (30, 2000),
            "Mobilya": (200, 20000),
            "Temizlik": (10, 200),
            "Kozmetik": (20, 1000),
            "Ofis": (2, 100),
            "Otomotiv": (20, 5000),
            "İnşaat": (10, 2000),
            "Bahçe": (15, 3000)
        }

        min_price, max_price = base_prices.get(category, (10, 1000))
        sale_price = round(random.uniform(min_price, max_price), 2)
        cost_price = round(sale_price * random.uniform(0.4, 0.7), 2)

        product = {
            "Ürün Kodu*": code,
            "Ürün Adı*": product_name,
            "Açıklama": f"{brand} marka {product_base.lower()} - yüksek kalite, uygun fiyat",
            "Barkod": generate_barcode(),
            "SKU": generate_sku(category, i),
            "Kategori Kodu": category,
            "Marka Kodu": brand,
            "Tedarikçi Kodu": f"TED-{str(random.randint(100, 999))}",
            "Birim*": random.choice(UNITS),
            "Ürün Tipi": random.choice(PRODUCT_TYPES),
            "Satış Fiyatı": sale_price,
            "Para Birimi": "TRY",
            "Maliyet Fiyatı": cost_price,
            "KDV Oranı (%)": random.choice([1, 8, 18, 20]),
            "Min Stok": random.randint(5, 50),
            "Max Stok": random.randint(100, 1000),
            "Yeniden Sipariş Seviyesi": random.randint(10, 30),
            "Yeniden Sipariş Miktarı": random.randint(20, 100),
            "Tedarik Süresi (Gün)": random.randint(1, 30),
            "Seri No Takibi": random.choice(["Evet", "Hayır"]),
            "Lot No Takibi": random.choice(["Evet", "Hayır"]),
            "Aktif": "Aktif"
        }
        products.append(product)

    return products

# ========================================
# Excel Writing
# ========================================

def create_excel_file(data, filename, sheet_name):
    """Create Excel file with formatted data"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not data:
        return

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    headers = list(data[0].keys())
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
            cell.border = thin_border

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(filename)
    print(f"Created: {filename} with {len(data)} records")

def main():
    """Main function to generate sample data files"""
    print("Generating sample data...")
    print("-" * 50)

    # Generate and save customer data
    customers = generate_customers(100)
    create_excel_file(
        customers,
        r"c:\Users\PC\source\repos\Stocker\customer_template.xlsx",
        "Müşteriler"
    )

    # Generate and save product data
    products = generate_products(100)
    create_excel_file(
        products,
        r"c:\Users\PC\source\repos\Stocker\product_template.xlsx",
        "Ürünler"
    )

    print("-" * 50)
    print("Sample data generation completed!")

if __name__ == "__main__":
    main()
